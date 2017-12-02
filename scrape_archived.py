import requests
from bs4 import BeautifulSoup
import sys
import urllib.parse
import json
import pyowm
from random import randint
# Change it to false to run the script completely
testing = True
# FileName of excel file
excelFileName = "Archived_ExcelFile.xlsx"

firstItemsEachCategory = []

baseURL = "https://trade.indiamart.com/buyersearch.mp?ss="

# listOfProxies = ['http://103.250.147.22:8080', 'http://103.192.64.10:8080', 'http://103.15.62.69:8080', 'http://103.219.192.147:9999', 'http://203.115.102.148:8080', 'http://110.173.183.50:80', 'http://182.74.200.207:80', 'http://110.173.183.63:80', 'http://103.205.15.129:8080', 'http://43.225.23.49:8080', 'http://110.173.183.57:80', 'http://117.202.20.66:555', 'http://27.106.125.21:8080', 'http://45.115.168.40:8080', 'http://111.119.210.10:8080', 'http://45.249.48.124:8080', 'http://45.124.145.34:8080', 'http://103.60.137.2:1', 'http://35.154.138.213:80', 'http://54.202.8.138:80', 'http://206.127.141.67:80', 'http://69.144.49.11:8080', 'http://45.77.132.79:33325', 'http://216.56.48.118:9000', 'http://54.177.186.237:80', 'http://162.243.138.193:80', 'http://47.89.241.103:3128', 'http://67.205.142.183:8080', 'http://47.88.32.46:3128', 'http://24.38.71.43:80', 'http://54.205.31.179:80', 'http://162.223.91.18:3128', 'http://47.88.84.190:8080', 'http://96.85.198.105:53281']
listOfProxies = ['http://143.0.188.8:80',
				 'http://139.59.47.12:3128',
				 'http://201.16.197.16:80',
				 'http://35.196.238.100:80',
				 'http://104.155.86.227:80',
				 'http://14.141.216.6:3128',
				 'http://118.69.140.108:53281',
				 'http://192.237.153.73:8080',
				 'http://45.77.153.228:55555',
				 'http://138.0.155.196:8080',
				 'http://203.74.4.2:80',
				 'http://203.74.4.7:80',
				 'http://35.195.221.39:80',
				 'http://45.6.216.66:3128',
				 'http://203.58.117.34:80',
				 'http://13.115.220.171:8080',
				 'http://87.98.157.128:3128',
				 'http://52.164.244.34:8080',
				 'http://45.76.88.240:3128',]
				 
def weather(place):
	owm = pyowm.OWM('8e47cb932d1448c4049c3506aca77f87')
	observation = owm.weather_at_place(place)
	w = observation.get_weather()
	complete_temp = w.get_temperature('celsius') 
	for i in complete_temp:
		if(i=="temp"):
			return complete_temp[i]

def getItems(categoryName, subCategoryName, subCategoryURL):
	data = readFirstItemFromFile()
	stopSubCat = False
	firstItem = True
	page_number = 0
	# while True:
		# if page_number == 0:
		# 	soup = BeautifulSoup(requests.get(subCategoryURL).content, "html.parser")
		# else:
		# 	if "No Buy Leads" in str(requests.get(subCategoryURL+"/buy"+str(page_number)+".html").content):
		# 		break
	x = urllib.parse.quote_plus(subCategoryName)
	while (True):
		try:
			indexOfProxy = randint(0, len(listOfProxies) - 1)
			prox = {
				'http' : listOfProxies[indexOfProxy],
				'https' : listOfProxies[indexOfProxy],
			}
			if useProxy:
				r = requests.get(baseURL+x, proxies=prox)
			else:
				r = requests.get(baseURL+x)
			if (r.status_code == 200):
				r = r.content
				break
		except requests.exceptions.Timeout:
			print ("Timeout Changing proxy", prox)
		except requests.exceptions.ProxyError:
			print ("Proxy not working", prox)
		except requests.exceptions.ConnectionError:
			print ("Proxy not working", prox)
	soup = BeautifulSoup(r, "html.parser")
	print ("[+] Getting items in " + categoryName + "->" + subCategoryName + str(page_number))
	itemDiv = soup.find_all("div", {"class": "trade-list"})
	# itemDiv = soup.find_all("div", {"class": "trade-list"})
	item_dictionary = dict()
	for div in itemDiv:
		location = "-"
		capacity = "-"
		quantity_unit = "-"
		quantity = "-"
		need_for_this = "-"
		frequency = "-"
		state = "-"
		country = "-"
		temperature = "-"

		itemName = div.find("div",{"class": "d_lm"}).find("p",{"class": "d_f1"}).find("a").text
		location = div.find("span", {"class": "bl_ccname location"}).text.lstrip().split(",")
		if len(location) == 1:
			country = location[0]
			temperature = weather(country)
		else:
			state = location[0]
			country = location[1]
			temperature = weather(state + ", " + country)

		date = div.find("span", {"class": "dtt updatedTime"}).text.lstrip()
		
		other_details = dict()

		other_details["Date"] = date
		other_details["Category"] = categoryName
		other_details["Sub Category"] = subCategoryName
		other_details["Item"] = itemName
		# other_details["Location"] = location
		other_details["State"] = state
		other_details["Country"] = country
		other_details["Temperature"] = temperature
		other_details["Capacity"] = "-"
		other_details["Quantity"] = "-"
		other_details["Quantity Unit"] = "-"
		other_details["Need/Usage"] = "-"
		other_details["Frequency"] = "-"

		flag_no_unit_check = False

		if div.find("div", {"class": "c15 pt4 fs pl"}) != None:
			for table_row in div.find("div", {"class": "c15 pt4 fs pl"}).find_all("tr"):
				data = table_row.text.replace("\n","").split(":")
				if "capacity" in data[0].lower():
					capacity = data[1].lstrip()
					other_details["Capacity"] = capacity
				if len(data[0].lower().split(" ")) == 2 and "unit" in data[0].lower():
					quantity_unit = data[1].lstrip()
					flag_no_unit_check = True
					other_details["Quantity Unit"] = quantity_unit
				if "quantity" in data[0].lower() and len(data[0].lower().split(" ")) != 2:
					quantity = data[1].lstrip().split(" ")[0]
					other_details["Quantity"] = quantity
					if not flag_no_unit_check and len(data[1].lstrip().split(" ")) > 1:
						quantity_unit = data[1].lstrip().split(" ")[1]
						other_details["Quantity Unit"] = quantity_unit
				if "need" in data[0].lower() or "usage" in data[0].lower():
					need_for_this = data[1].lstrip()
					other_details["Need/Usage"] = need_for_this
				if "frequency" in data[0].lower():
					frequency = data[1].lstrip()
					other_details["Frequency"] = frequency
				item_dictionary[itemName] = other_details
			if firstItem:
				firstItemsEachCategory.append(itemName)
				firstItem = False
			else:
				if data != []:
					for i in data:
						if itemName == i:
							stopSubCat = True
							break
		if stopSubCat:
			break

	# if testing:
	# 	break
	return item_dictionary
		
def writeToExcel(itemNumber, dictToWrite, row, worksheet):
	# worksheet.write(row, 1, itemNumber)
	ws.cell(row = row, column = 1).value = itemNumber
	j = 2
	for i in dictToWrite.values():
		ws.cell(row = row, column = j).value = i
		j += 1

def writeHeadingToExcel(worksheet):
	heading = ['S.No', 'Date', 'Category', 'Sub Category', 'Item', 'State', 'Country', 'Temperature', 'Capacity', 'Quantity', 'Quantity Unit', 'Need for this/usage', 'Frequency']
	j = 1
	for i in heading:
		ws.cell(row = 1, column = j).value = i
		j += 1

def writeFirstItemToFile(firstItemList):
	import pickle
	with open("firstItemHelper.txt", "wb") as fp:
		pickle.dump(firstItemList, fp)

def readFirstItemFromFile():
	import pickle
	data = []
	try:
		with open("firstItemHelper.txt", "wb") as fp:
			data = pickle.load(fp)
	except:
		pass
	return data

if __name__ == '__main__':
	subCategories = {}
	with open('subcategories.txt', 'r') as f:
		l = f.readlines()
		s = l[0]
		json_acceptable_string = s.replace("'", "\"")
		subCategories = json.loads(json_acceptable_string)

	# for i in subCategories:
	# 	for j in subCategories[i]:
	# 		toWrite = getItems(i, i, j)
	# 		print (toWrite)


	itemNumber = 1
	from openpyxl import Workbook
	from openpyxl import load_workbook
	wb = Workbook()
	ws = wb.active
	# workbook = xlsxwriter.Workbook(excelFileName)
	# worksheet = workbook.add_worksheet()
	writeHeadingToExcel(ws)
	wb.save(excelFileName)
	try:
		for i in subCategories:
			for j in subCategories[i]:
				dictToWrite = getItems(i, i, j)
				newDictionary = dict()
				for j in dictToWrite:
					wb = load_workbook(filename = excelFileName)
					ws = wb.get_active_sheet()
					# print(dictToWrite[j])
					writeToExcel(itemNumber, dictToWrite[j], itemNumber + 1, ws)
					wb.save(excelFileName)
					itemNumber += 1
					print(itemNumber)
				if testing:
					break
	except KeyboardInterrupt:
		pass
	# except:
	# 	 print ("Unexpected error:", sys.exc_info()[0])
	wb = load_workbook(filename = excelFileName)
	ws = wb.get_active_sheet()
	writeFirstItemToFile(firstItemsEachCategory)
	wb.save(excelFileName)
