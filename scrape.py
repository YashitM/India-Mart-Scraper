import requests
from bs4 import BeautifulSoup
from random import randint

# Change it to false to run the script completely
testing = True
# FileName of excel file
excelFileName = "ExcelFile.xlsx"

firstItemsEachCategory = []

listOfProxies = ['http://103.250.147.22:8080', 'http://103.192.64.10:8080', 'http://103.15.62.69:8080', 'http://103.219.192.147:9999', 'http://203.115.102.148:8080', 'http://110.173.183.50:80', 'http://182.74.200.207:80', 'http://110.173.183.63:80', 'http://103.205.15.129:8080', 'http://43.225.23.49:8080', 'http://110.173.183.57:80', 'http://117.202.20.66:555', 'http://27.106.125.21:8080', 'http://45.115.168.40:8080', 'http://111.119.210.10:8080', 'http://45.249.48.124:8080', 'http://45.124.145.34:8080', 'http://103.60.137.2:1', 'http://35.154.138.213:80', 'http://54.202.8.138:80', 'http://206.127.141.67:80', 'http://69.144.49.11:8080', 'http://45.77.132.79:33325', 'http://216.56.48.118:9000', 'http://54.177.186.237:80', 'http://162.243.138.193:80', 'http://47.89.241.103:3128', 'http://67.205.142.183:8080', 'http://47.88.32.46:3128', 'http://24.38.71.43:80', 'http://54.205.31.179:80', 'http://162.223.91.18:3128', 'http://47.88.84.190:8080', 'http://96.85.198.105:53281']

def weather(place):
	owm = pyowm.OWM('8e47cb932d1448c4049c3506aca77f87')
	observation = owm.weather_at_place(place)
	w = observation.get_weather()
	complete_temp = w.get_temperature('celsius') 
	for i in complete_temp:
		if(i=="temp"):
			return complete_temp[i]

def getCategories():
	while (True):
		try:
			indexOfProxy = randint(0, len(listOfProxies) - 1)
			prox = {
				'http' : listOfProxies[indexOfProxy],
				'https' : listOfProxies[indexOfProxy],
			}
			r = requests.get("https://trade.indiamart.com", proxies=prox)
			if (r.status_code == 200):
				r = r.content
				break
		except requests.exceptions.Timeout:
			print ("Timeout Changing proxy", prox)
		except requests.exceptions.ProxyError:
			print ("Proxy not working", prox)
	soup = BeautifulSoup(r, "html.parser")
	categories_dictionary = dict()

	for item in soup.find_all("li", {"class": "dc-mega-li"}):
		if item.find("a").get("href") != "#":
			categories_dictionary[item.text.lstrip().rstrip("\n")] = item.find("a").get("href")
	# do the same for 132 and 133
	for item in soup.find_all("li", {"class": "menu-item-131"}):
		categories_dictionary[item.text.lstrip().rstrip("\n")] = item.find("a").get("href")

	print("[*] Categories found!")
	return categories_dictionary

def getSubcategories(categoryName, categoryURL):
	while (True):
		try:
			indexOfProxy = randint(0, len(listOfProxies) - 1)
			prox = {
				'http' : listOfProxies[indexOfProxy],
				'https' : listOfProxies[indexOfProxy],
			}
			r = requests.get(categoryURL, proxies=prox)
			if (r.status_code == 200):
				r = r.content
				break
		except requests.exceptions.Timeout:
			print ("Timeout Changing proxy", prox)
		except requests.exceptions.ProxyError:
			print ("Proxy not working", prox)
	soup = BeautifulSoup(r, "html.parser")
	print ("[+] Finding Sub Categories of " + categoryName)
	subCategories_dictionary = dict()

	for ul in soup.find_all("ul", {"class": "grouplist"}):
		for li in ul.find_all("li"):
			if li.find("a").get("href") != "#":
				text = li.text.lstrip().rstrip("\n")
				subCategories_dictionary[text[:text.find("(")]] = li.find("a").get("href")

	return subCategories_dictionary

def getItems(categoryName, subCategoryName, subCategoryURL):
	data = readFirstItemFromFile()
	stopSubCat = False
	firstItem = True
	page_number = 0
	while True:
		if page_number == 0:
			while (True):
				try:
					indexOfProxy = randint(0, len(listOfProxies) - 1)
					prox = {
						'http' : listOfProxies[indexOfProxy],
						'https' : listOfProxies[indexOfProxy],
					}
					r = requests.get(subCategoryURL, proxies=prox)
					if (r.status_code == 200):
						r = r.content
						break
				except requests.exceptions.Timeout:
					print ("Timeout Changing proxy", prox)
				except requests.exceptions.ProxyError:
					print ("Proxy not working", prox)
			soup = BeautifulSoup(r, "html.parser")
		else:
			if "No Buy Leads" in requests.get(subCategoryURL+"/buy"+str(page_number)+".html").content:
				break
		print ("[+] Getting items in " + categoryName + "->" + subCategoryName + str(page_number))
		itemDiv = soup.find_all("div", {"class": "blockdiv"})
		# itemDiv = soup.find_all("div", {"class": "trade-list"})
		item_dictionary = dict()
		for div in itemDiv:
			location = "-"
			capacity = "-"
			quantity_unit = "-"
			quantity = "-"
			need_for_this = "-"
			frequency = "-"
			state = "-"
			country = "-"
			temperature = "-"

			itemName = div.find("div",{"class": "d_lm"}).find("p",{"class": "d_f1"}).find("a").text
			location = div.find("span", {"class": "bl_ccname location"}).text.lstrip().split(",")
			if len(location) == 1:
				country = location[0]
				temperature = weather(country)
			else:
				state = location[0]
				country = location[1]
				temperature = weather(state + ", " + country)

			date = div.find("span", {"class": "dtt updatedTime"}).text.lstrip()
			
			other_details = dict()

			other_details["Date"] = date
			other_details["Category"] = categoryName
			other_details["Sub Category"] = subCategoryName
			other_details["Item"] = itemName
			# other_details["Location"] = location
			other_details["State"] = state
			other_details["Country"] = country
			other_details["Temperature"] = temperature
			other_details["Capacity"] = "-"
			other_details["Quantity"] = "-"
			other_details["Quantity Unit"] = "-"
			other_details["Need/Usage"] = "-"
			other_details["Frequency"] = "-"

			flag_no_unit_check = False

			if div.find("div", {"class": "c15 pt4 fs pl"}) != None:
				for table_row in div.find("div", {"class": "c15 pt4 fs pl"}).find_all("tr"):
					data = table_row.text.replace("\n","").split(":")
					if "capacity" in data[0].lower():
						capacity = data[1].lstrip()
						other_details["Capacity"] = capacity
					if len(data[0].lower().split(" ")) == 2 and "unit" in data[0].lower():
						quantity_unit = data[1].lstrip()
						flag_no_unit_check = True
						other_details["Quantity Unit"] = quantity_unit
					if "quantity" in data[0].lower() and len(data[0].lower().split(" ")) != 2:
						quantity = data[1].lstrip().split(" ")[0]
						other_details["Quantity"] = quantity
						if not flag_no_unit_check and len(data[1].lstrip().split(" ")) > 1:
							quantity_unit = data[1].lstrip().split(" ")[1]
							other_details["Quantity Unit"] = quantity_unit
					if "need" in data[0].lower() or "usage" in data[0].lower():
						need_for_this = data[1].lstrip()
						other_details["Need/Usage"] = need_for_this
					if "frequency" in data[0].lower():
						frequency = data[1].lstrip()
						other_details["Frequency"] = frequency
					item_dictionary[itemName] = other_details
				if firstItem:
					firstItemsEachCategory.append(itemName)
					firstItem = False
				else:
					if data != []:
						for i in data:
							if itemName == i:
								stopSubCat = True
								break
		if stopSubCat:
			break

		page_number+=1
		if testing:
			break
	return item_dictionary
		
def writeToExcel(itemNumber, dictToWrite, row, worksheet):
	# worksheet.write(row, 1, itemNumber)
	ws.cell(row = row, column = 1).value = itemNumber
	j = 2
	for i in dictToWrite.values():
		ws.cell(row = row, column = j).value = i
		j += 1

def writeHeadingToExcel(worksheet):
	heading = ['S.No', 'Date', 'Category', 'Sub Category', 'Item', 'State', 'Country', 'Temperature', 'Capacity', 'Quantity', 'Quantity Unit', 'Need for this/usage', 'Frequency']
	j = 1
	for i in heading:
		ws.cell(row = 1, column = j).value = i
		j += 1

def writeFirstItemToFile(firstItemList):
	import pickle
	with open("firstItemHelper.txt", "wb") as fp:
		pickle.dump(firstItemList, fp)

def readFirstItemFromFile():
	import pickle
	data = []
	try:
		with open("firstItemHelper.txt", "wb") as fp:
			data = pickle.load(fp)
	except:
		pass
	return data

if __name__ == '__main__':
	categories = getCategories()
	subCategories = dict()
	j = 0
	for i in categories:
		subCategories[i] = getSubcategories(i, categories[i])
		if testing:
			break



	itemNumber = 1
	from openpyxl import Workbook
	wb = Workbook()
	ws = wb.active
	# workbook = xlsxwriter.Workbook(excelFileName)
	# worksheet = workbook.add_worksheet()
	writeHeadingToExcel(ws)
	try:
		for i in subCategories:
			for j in subCategories[i]:
				dictToWrite = getItems(i,j,subCategories[i][j])
				newDictionary = dict()
				for j in dictToWrite:
					# print(dictToWrite[j])
					writeToExcel(itemNumber, dictToWrite[j], itemNumber + 1, ws)
					itemNumber += 1
					print(itemNumber)
				if testing:
					break
	except KeyboardInterrupt:
		pass
	except:
		 print ("Unexpected error:", sys.exc_info()[0])
	writeFirstItemToFile(firstItemsEachCategory)
	wb.save(excelFileName)
